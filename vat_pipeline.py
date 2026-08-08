"""
Reusable VAT return pipeline for Knights Telecom (v2).

Produces SEPARATE files per deliverable, all saved into the period's output
folder (e.g. "VAT - June'26"):
    Knights_Telecom_VAT_Return_<Month>_<Year>.xlsx   - return + schedules
    Invoice_Series_Check_<Month>_<Year>.xlsx         - gap check on invoice numbers
    Ledger_Output_VAT_<Month>_<Year>.xlsx
    Ledger_Input_VAT_<Month>_<Year>.xlsx
    Ledger_Sales_<Month>_<Year>.xlsx
    Trial_Balance_<Month>_<Year>.xlsx
    Journal_Entry_<Month>_<Year>.xlsx                - post-filing, human confirms before posting

Every data range is a real Excel Table (banded rows, filter buttons, borders
built in) rather than manually styled cells. Columns are auto-sized to their
content, header rows are frozen.
"""
import glob
import os
import re
import subprocess
from datetime import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

import zoho_books_api

RECALC_SCRIPT = os.environ.get('VAT_RECALC_SCRIPT', '/mnt/skills/public/xlsx/scripts/recalc.py')
CYBERKNIGHT_NAMES = {'Cyber Knight Technologies FZ LLC', 'Cyber Knight Technologies FZ-LLC'}

TITLE_FONT = Font(name='Arial', bold=True, size=14)
BOLD_FONT = Font(name='Arial', bold=True, size=10)
NORMAL_FONT = Font(name='Arial', size=10)
NOTE_FONT = Font(name='Arial', italic=True, size=9, color='808080')
WARN_FONT = Font(name='Arial', italic=True, size=9, color='C00000')
TOTAL_FILL = PatternFill('solid', fgColor='D9E1F2')
MONEY = '#,##0.00'
PCT = '0%'
DATE_FMT = 'DD-MMM-YYYY'

_table_counter = 0


def to_float(v):
    if v is None or v == '':
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _find_one(source_dir, pattern):
    matches = glob.glob(os.path.join(source_dir, pattern))
    if not matches:
        raise FileNotFoundError(f"No file matching {pattern} in {source_dir}")
    return sorted(matches)[-1]


def rows_of(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    return idx, list(ws.iter_rows(min_row=2, values_only=True))


def add_table(ws, start_row, headers, data_rows, start_col=1, title=None, title_row=None):
    """
    Writes `headers` at start_row and `data_rows` beneath it, formats the
    range as a real Excel Table (banded rows + filter buttons + borders come
    from the table style, not manual cell formatting), auto-sizes columns,
    and freezes the header row. Returns the row index of the first empty
    row after the table.
    """
    global _table_counter
    ncols = len(headers)
    if title:
        ws.cell(row=title_row or (start_row - 2), column=start_col, value=title).font = TITLE_FONT

    for c, h in enumerate(headers):
        ws.cell(row=start_row, column=start_col + c, value=h)
    ws.row_dimensions[start_row].height = 30

    r = start_row + 1
    col_widths = [len(str(h)) + 2 for h in headers]
    for row_vals in data_rows:
        for c, v in enumerate(row_vals):
            cell = ws.cell(row=r, column=start_col + c, value=v)
            cell.font = NORMAL_FONT
            col_widths[c] = max(col_widths[c], len(str(v)) if v is not None else 0)
        r += 1
    last_row = r - 1

    if last_row >= start_row + 1:
        first_col_letter = get_column_letter(start_col)
        last_col_letter = get_column_letter(start_col + ncols - 1)
        _table_counter += 1
        table_name = f"Tbl{_table_counter}_{re.sub(r'[^A-Za-z0-9]', '', ws.title)[:15]}"
        tab = Table(displayName=table_name, ref=f"{first_col_letter}{start_row}:{last_col_letter}{last_row}")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(tab)
        ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col).coordinate

    for c, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(start_col + c)].width = min(max(w + 2, 10), 55)

    return last_row + 1


def apply_money_cols(ws, start_row, end_row, cols, fmt=MONEY):
    for r in range(start_row, end_row + 1):
        for c in cols:
            ws.cell(row=r, column=c).number_format = fmt


def apply_date_cols(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in cols:
            ws.cell(row=r, column=c).number_format = DATE_FMT


def write_total_row(ws, row, label_col, label, sum_cols, first_data_row, last_data_row):
    ws.cell(row=row, column=label_col, value=label).font = BOLD_FONT
    for c in sum_cols:
        letter = get_column_letter(c)
        cell = ws.cell(row=row, column=c, value=f"=SUM({letter}{first_data_row}:{letter}{last_data_row})")
        cell.font = BOLD_FONT
        cell.number_format = MONEY
        cell.fill = TOTAL_FILL
    ws.cell(row=row, column=label_col).fill = TOTAL_FILL


# ---------------- Classification (now reads Zoho API record dicts) ---------
def extract_sales(invoice_records):
    std = defaultdict(lambda: {'inv_date': None, 'supply_date': None, 'customer': '', 'trn': '',
                                'item_code': set(), 'desc': [], 'taxable': 0.0, 'vat': 0.0,
                                'sub_total': 0.0, 'total': 0.0})
    zero_export = defaultdict(lambda: {'inv_date': None, 'customer': '', 'trn': '', 'taxable': 0.0, 'vat': 0.0,
                                        'sub_total': 0.0, 'total': 0.0})
    for inv in invoice_records:
        inv_no = inv.get('invoice_number')
        if not inv_no:
            continue
        # verify: field name for the customer's VAT/tax registration number —
        # Zoho commonly exposes this as 'tax_reg_no' or 'gst_no' depending on
        # edition; adjust if your org's response uses a different key.
        trn = inv.get('tax_reg_no') or inv.get('gst_no') or ''
        vat_treatment = inv.get('vat_treatment', '')
        target = std if vat_treatment == 'vat_registered' else zero_export
        d = target[inv_no]
        d['inv_date'] = inv.get('date')
        if 'supply_date' in d:
            # verify: no standard "date of supply" field exists on a Zoho invoice
            # distinct from the invoice date unless tracked via a custom field.
            # Falls back to the invoice date; swap in the custom field key if
            # Knights Telecom tracks supply date separately.
            d['supply_date'] = inv.get('date')
        d['customer'] = inv.get('customer_name', '')
        d['trn'] = trn
        d['sub_total'] = to_float(inv.get('sub_total'))
        d['total'] = to_float(inv.get('total'))
        for li in inv.get('line_items', []):
            item_total = to_float(li.get('item_total'))
            item_tax = to_float(li.get('tax_amount'))
            if vat_treatment == 'vat_registered':
                item_code = li.get('sku') or li.get('name')
                if item_code:
                    d['item_code'].add(str(item_code))
                desc = li.get('description') or li.get('name') or ''
                if desc:
                    d['desc'].append(str(desc).split('\n')[0])
            d['taxable'] += item_total
            d['vat'] += item_tax
    return std, zero_export


def extract_credit_notes(cn_records):
    cns = defaultdict(lambda: {'cn_date': None, 'orig_inv': set(), 'customer': '', 'trn': '',
                                'desc': [], 'taxable': 0.0, 'vat': 0.0, 'sub_total': 0.0, 'total': 0.0})
    for cn in cn_records:
        cn_no = cn.get('creditnote_number')
        if not cn_no or cn.get('vat_treatment') != 'vat_registered':
            continue
        d = cns[cn_no]
        d['cn_date'] = cn.get('date')
        # verify: field name linking a credit note to its original invoice —
        # commonly 'invoices_credited' (a list) or 'reference_invoice_id' on
        # some editions.
        for linked in cn.get('invoices_credited', []):
            num = linked.get('invoice_number')
            if num:
                d['orig_inv'].add(str(num))
        d['customer'] = cn.get('customer_name', '')
        d['trn'] = cn.get('tax_reg_no') or cn.get('gst_no') or ''
        d['sub_total'] = to_float(cn.get('sub_total'))
        d['total'] = to_float(cn.get('total'))
        for li in cn.get('line_items', []):
            desc = li.get('description') or li.get('name') or ''
            if desc:
                d['desc'].append(str(desc).split('\n')[0])
            d['taxable'] += to_float(li.get('item_total'))
            d['vat'] += to_float(li.get('tax_amount'))
    return cns


def extract_bills(bill_records):
    purchases = defaultdict(lambda: {'bill_date': None, 'vendor': '', 'trn': '', 'taxable': 0.0, 'vat': 0.0,
                                      'desc': [], 'sub_total': 0.0, 'total': 0.0})
    rcm = defaultdict(lambda: {'bill_date': None, 'vendor': '', 'taxable': 0.0, 'desc': [],
                                'sub_total': 0.0, 'total': 0.0})
    for bill in bill_records:
        bill_no = bill.get('bill_number')
        if not bill_no:
            continue
        vendor = (bill.get('vendor_name') or '').strip()
        vat_treatment = bill.get('vat_treatment', '')
        sub_total = to_float(bill.get('sub_total'))
        total = to_float(bill.get('total'))
        line_items = bill.get('line_items', [])
        item_total_sum = sum(to_float(li.get('item_total')) for li in line_items)
        # verify: some editions report tax at the bill level under a "taxes"
        # array (like Expenses do) rather than per line-item "tax_amount" —
        # check your org's response; this assumes per-line tax_amount exists.
        tax_amt_sum = sum(to_float(li.get('tax_amount')) for li in line_items)
        descs = []
        for li in line_items:
            desc = li.get('description') or li.get('name') or ''
            if desc:
                descs.append(str(desc).split('\n')[0])

        if vendor in CYBERKNIGHT_NAMES:
            d = rcm[bill_no]
            d['bill_date'] = bill.get('date')
            d['vendor'] = vendor
            d['desc'] = descs
            d['taxable'] += item_total_sum
            d['sub_total'] = sub_total
            d['total'] = total
        elif vat_treatment == 'vat_registered' and tax_amt_sum > 0:
            d = purchases[bill_no]
            d['bill_date'] = bill.get('date')
            d['vendor'] = vendor
            d['trn'] = bill.get('tax_reg_no') or bill.get('gst_no') or ''
            d['desc'] = descs
            d['taxable'] += item_total_sum
            d['vat'] += tax_amt_sum
            d['sub_total'] = sub_total
            d['total'] = total
    return purchases, rcm


def extract_expenses(expense_records):
    purchases = defaultdict(lambda: {'exp_date': None, 'vendor': '', 'trn': '', 'taxable': 0.0, 'vat': 0.0,
                                      'desc': [], 'sub_total': 0.0, 'total': 0.0})
    for i, exp in enumerate(expense_records):
        # verify: Expenses report tax via a "taxes": [{"tax_id":..,"tax_amount":..}]
        # array rather than a flat tax_amount field — sum it here.
        tax_amt = sum(to_float(t.get('tax_amount')) for t in exp.get('taxes', []))
        vat_treatment = exp.get('vat_treatment', '')
        if vat_treatment == 'vat_registered' and tax_amt > 0:
            key = f"EXP-{i+1:04d}"
            d = purchases[key]
            d['exp_date'] = exp.get('date')
            d['vendor'] = exp.get('vendor_name') or exp.get('paid_through_account_name') or ''
            d['trn'] = exp.get('tax_reg_no') or exp.get('gst_no') or ''
            d['desc'].append(str(exp.get('description') or exp.get('account_name') or ''))
            amount = to_float(exp.get('amount'))
            d['taxable'] += amount
            d['vat'] += tax_amt
            d['sub_total'] = amount
            d['total'] = amount + tax_amt
    return purchases


# ---------------- File 1: VAT Return + schedules ----------------------------
def build_return_workbook(period_label, std_sales, zero_export, cns, purch_bills, rcm, purch_exp):
    global _table_counter
    _table_counter = 0
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sales 15%
    ws = wb.create_sheet('Sales 15%')
    sales_rows = []
    for inv_no, d in sorted(std_sales.items(), key=lambda x: (x[1]['inv_date'] or datetime.min)):
        sales_rows.append([inv_no, d['inv_date'], d.get('supply_date'), ', '.join(sorted(d['item_code']))[:100],
                            '; '.join(d['desc'][:3]), d['trn'], d['customer'], 0.15,
                            round(d['taxable'], 2), round(d['vat'], 2)])
    headers = ['Tax Invoice No', 'Tax Invoice Date', 'Date of Supply', 'Item Code', 'Description',
               'VAT No of Customer', 'Name of Customer', 'Tax Rate', 'Taxable Amount', 'Amount of VAT']
    next_row = add_table(ws, 4, headers, sales_rows, title=f'STANDARD RATED SALES - {period_label}', title_row=2)
    apply_date_cols(ws, 5, next_row - 1, [2, 3])
    apply_money_cols(ws, 5, next_row - 1, [9, 10])
    for r in range(5, next_row):
        ws.cell(row=r, column=8).number_format = PCT
    SALES_TOTAL_ROW = next_row + 1
    write_total_row(ws, SALES_TOTAL_ROW, 1, 'Total', [9, 10], 5, next_row - 1)

    zrow0 = SALES_TOTAL_ROW + 3
    zero_rows = []
    for inv_no, d in sorted(zero_export.items()):
        zero_rows.append([inv_no, d['inv_date'], d['customer'], d['trn'], round(d['taxable'], 2), round(d['vat'], 2)])
    zheaders = ['Invoice No', 'Invoice Date', 'Customer', 'TRN', 'Taxable Amount', 'VAT']
    znext = add_table(ws, zrow0 + 1, zheaders, zero_rows,
                       title='NON-GCC / ZERO-RATED SALES (review only, excluded from Standard Rated total)', title_row=zrow0)
    apply_date_cols(ws, zrow0 + 2, znext - 1, [2])
    apply_money_cols(ws, zrow0 + 2, znext - 1, [5, 6])
    ZERO_TOTAL_ROW = znext + 1
    write_total_row(ws, ZERO_TOTAL_ROW, 1, 'Total', [5], zrow0 + 2, znext - 1)

    # Credit Notes 15%
    ws = wb.create_sheet('Credit Notes 15%')
    cn_rows = []
    for cn_no, d in sorted(cns.items(), key=lambda x: (x[1]['cn_date'] or datetime.min)):
        cn_rows.append([cn_no, d['cn_date'], ', '.join(sorted(d['orig_inv'])), '; '.join(d['desc'][:3]),
                         d['trn'], d['customer'], 0.15, round(d['taxable'], 2), round(d['vat'], 2)])
    headers = ['Credit Note No', 'CN Date', 'Corresponding Invoice', 'Description', 'VAT No of Customer',
               'Name of Customer', 'Tax Rate', 'Taxable Amount', 'Amount of VAT']
    next_row = add_table(ws, 4, headers, cn_rows, title=f'CREDIT NOTES ISSUED - {period_label}', title_row=2)
    apply_date_cols(ws, 5, next_row - 1, [2])
    apply_money_cols(ws, 5, next_row - 1, [8, 9])
    for r in range(5, next_row):
        ws.cell(row=r, column=7).number_format = PCT
    CN_TOTAL_ROW = next_row + 1
    write_total_row(ws, CN_TOTAL_ROW, 1, 'Total', [8, 9], 5, next_row - 1)

    # Purchase 15% (bills + expenses)
    ws = wb.create_sheet('Purchase 15%')
    combined = [('Bill', k, d['bill_date'], '; '.join(d['desc'][:2]), d['trn'], d['vendor'], d['taxable'], d['vat'])
                for k, d in purch_bills.items()]
    combined += [('Expense', k, d['exp_date'], '; '.join(d['desc'][:2]), d['trn'], d['vendor'], d['taxable'], d['vat'])
                 for k, d in purch_exp.items()]
    combined.sort(key=lambda x: (x[2] or datetime.min))
    purch_rows = [[src, ref, dt, desc, trn, vendor, 0.15, round(tax, 2), round(vat, 2)]
                  for src, ref, dt, desc, trn, vendor, tax, vat in combined]
    headers = ['Source', 'Reference No', 'Date', 'Description', 'VAT No of Supplier', 'Name of Supplier',
               'Tax Rate', 'Taxable Amount', 'Amount of VAT']
    next_row = add_table(ws, 4, headers, purch_rows, title=f'STANDARD RATED PURCHASES - {period_label}', title_row=2)
    apply_date_cols(ws, 5, next_row - 1, [3])
    apply_money_cols(ws, 5, next_row - 1, [8, 9])
    for r in range(5, next_row):
        ws.cell(row=r, column=7).number_format = PCT
    PURCH_TOTAL_ROW = next_row + 1
    write_total_row(ws, PURCH_TOTAL_ROW, 1, 'Total', [8, 9], 5, next_row - 1)

    # Import of services - RCM
    ws = wb.create_sheet('Import of services-RCM')
    rcm_rows = []
    for bill_no, d in sorted(rcm.items(), key=lambda x: (x[1]['bill_date'] or datetime.min)):
        value = round(d['taxable'], 2)
        rcm_rows.append([bill_no, d['bill_date'], '; '.join(d['desc'][:2]), d['vendor'], 0.15, value,
                          round(value * 0.15, 2), round(value * 0.15, 2)])
    headers = ['Supplier Invoice No', 'Date', 'Description', 'Name of Supplier', 'Tax Rate', 'Value of Service',
               'Notional VAT (Output)', 'Notional VAT (Input)']
    next_row = add_table(ws, 4, headers, rcm_rows, title=f'IMPORTS OF SERVICES VIA RCM - {period_label}', title_row=2)
    apply_date_cols(ws, 5, next_row - 1, [2])
    apply_money_cols(ws, 5, next_row - 1, [6, 7, 8])
    for r in range(5, next_row):
        ws.cell(row=r, column=5).number_format = PCT
    RCM_TOTAL_ROW = next_row + 1
    write_total_row(ws, RCM_TOTAL_ROW, 1, 'Total', [6, 7, 8], 5, next_row - 1)

    # VAT Return computation (not a table — this is a structured form, keep as styled cells)
    ws = wb.create_sheet('VAT Return', 0)
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 3
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 3
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 3
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 3
    ws.column_dimensions['I'].width = 18
    ws['B3'] = 'Knights Telecom'; ws['B3'].font = TITLE_FONT
    ws['B4'] = 'Company Name: Knights Telecom'; ws['B4'].font = BOLD_FONT
    ws['B5'] = f'Period: {period_label}'; ws['B5'].font = BOLD_FONT
    ws['B6'] = 'VAT Return Computation (auto-generated from books)'; ws['B6'].font = BOLD_FONT

    r = 9
    ws.cell(row=r, column=2, value='Particulars').font = BOLD_FONT
    ws.cell(row=r, column=5, value='Total Amount').font = BOLD_FONT
    ws.cell(row=r, column=7, value='Adjustments').font = BOLD_FONT
    ws.cell(row=r, column=9, value='VAT').font = BOLD_FONT
    r = 12
    sales_rows_start = r
    rows_spec = [
        ('1', 'Standard Rated sales - 15%', f"='Sales 15%'!I{SALES_TOTAL_ROW}-'Credit Notes 15%'!H{CN_TOTAL_ROW}", 0,
         f"='Sales 15%'!J{SALES_TOTAL_ROW}-'Credit Notes 15%'!I{CN_TOTAL_ROW}"),
        ('1.2', 'Standard Rated sales - 5%', 0, 0, 0),
        ('2', 'Private Healthcare / Private Education', 0, 0, 0),
        ('3', 'Zero rated domestic Sales', 0, 0, 0),
        ('4', 'Exports / Non-GCC sales', f"='Sales 15%'!E{ZERO_TOTAL_ROW}", 0, 0),
        ('5', 'Exempt Sales', 0, 0, 0),
    ]
    for num, label, amt, adj, vat in rows_spec:
        ws.cell(row=r, column=2, value=num)
        ws.cell(row=r, column=3, value=label)
        ws.cell(row=r, column=5, value=amt).number_format = MONEY
        ws.cell(row=r, column=7, value=adj).number_format = MONEY
        ws.cell(row=r, column=9, value=vat).number_format = MONEY
        r += 1
    sales_rows_end = r - 1
    r += 1
    ws.cell(row=r, column=3, value='Total Sales').font = BOLD_FONT
    ws.cell(row=r, column=5, value=f"=SUM(E{sales_rows_start}:E{sales_rows_end})").number_format = MONEY
    ws.cell(row=r, column=9, value=f"=SUM(I{sales_rows_start}:I{sales_rows_end})").number_format = MONEY
    for c in (5, 7, 9):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
    TOTAL_SALES_ROW = r
    r += 2
    purch_rows_start = r
    purch_spec = [
        ('7', 'Standard rated domestic purchases - 15%', f"='Purchase 15%'!H{PURCH_TOTAL_ROW}", 0, f"='Purchase 15%'!I{PURCH_TOTAL_ROW}"),
        ('7.2', 'Standard rated domestic purchases - 5%', 0, 0, 0),
        ('8', 'Imports subject to VAT paid at Customs - 15%', 0, 0, 0),
        ('8.2', 'Imports subject to VAT paid at Customs - 5%', 0, 0, 0),
        ('9', 'Imports via reverse charge mechanism - 15%', f"='Import of services-RCM'!F{RCM_TOTAL_ROW}", 0, f"='Import of services-RCM'!H{RCM_TOTAL_ROW}"),
        ('9.2', 'Imports via reverse charge mechanism - 5%', 0, 0, 0),
        ('10', 'Zero Rated Purchases', 0, 0, 0),
        ('11', 'Exempt Purchases', 0, 0, 0),
    ]
    for num, label, amt, adj, vat in purch_spec:
        ws.cell(row=r, column=2, value=num)
        ws.cell(row=r, column=3, value=label)
        ws.cell(row=r, column=5, value=amt).number_format = MONEY
        ws.cell(row=r, column=7, value=adj).number_format = MONEY
        ws.cell(row=r, column=9, value=vat).number_format = MONEY
        r += 1
    purch_rows_end = r - 1
    r += 1
    ws.cell(row=r, column=3, value='Total Purchases').font = BOLD_FONT
    ws.cell(row=r, column=5, value=f"=SUM(E{purch_rows_start}:E{purch_rows_end})").number_format = MONEY
    ws.cell(row=r, column=9, value=f"=SUM(I{purch_rows_start}:I{purch_rows_end})").number_format = MONEY
    for c in (5, 7, 9):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
    TOTAL_PURCH_ROW = r
    r += 2
    ws.cell(row=r, column=3, value='Total VAT due for current period').font = BOLD_FONT
    rcm_vat_cell = f"I{purch_rows_start + 4}"
    ws.cell(row=r, column=9, value=f"=I{TOTAL_SALES_ROW}+{rcm_vat_cell}-I{TOTAL_PURCH_ROW}").number_format = MONEY
    NET_VAT_ROW = r
    r += 1
    ws.cell(row=r, column=3, value='(RCM self-charged output tax added back — net RCM impact is nil)').font = NOTE_FONT
    r += 2
    ws.cell(row=r, column=3, value='Corrections from previous period (+/- SAR 5,000)')
    ws.cell(row=r, column=9, value=0).number_format = MONEY
    CORR_ROW = r
    r += 2
    ws.cell(row=r, column=3, value='VAT credit carried forward')
    ws.cell(row=r, column=9, value=0).number_format = MONEY
    CF_ROW = r
    r += 2
    ws.cell(row=r, column=3, value='Net VAT due (or claim)').font = BOLD_FONT
    ws.cell(row=r, column=9, value=f"=I{NET_VAT_ROW}+I{CORR_ROW}-I{CF_ROW}").number_format = MONEY
    ws.cell(row=r, column=9).fill = TOTAL_FILL
    FINAL_ROW = r

    return wb, FINAL_ROW


# ---------------- File 2: Invoice Series Check -------------------------------
def build_invoice_series_workbook(period_label, std_sales, zero_export):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Invoice Series Check'

    all_invoices = list(std_sales.keys()) + list(zero_export.keys())
    parsed = []
    unparsed = []
    for inv in all_invoices:
        m = re.search(r'(\d+)$', str(inv))
        if m:
            parsed.append((int(m.group(1)), inv))
        else:
            unparsed.append(inv)
    parsed.sort()

    rows = []
    gaps_found = 0
    for i, (num, inv) in enumerate(parsed):
        if i == 0:
            diff = None
            flag = ''
        else:
            diff = num - parsed[i - 1][0]
            flag = 'OK' if diff == 1 else f'GAP of {diff - 1}'
            if diff != 1:
                gaps_found += 1
        rows.append([inv, num, diff, flag])

    headers = ['Invoice No', 'Numeric Sequence', 'Diff from Previous', 'Check']
    add_table(ws, 4, headers, rows, title=f'INVOICE SEQUENTIAL SERIES CHECK - {period_label}', title_row=2)

    summary_row = 4 + len(rows) + 3
    ws.cell(row=summary_row, column=1, value='Summary').font = BOLD_FONT
    ws.cell(row=summary_row + 1, column=1, value=f'Total invoices checked: {len(parsed)}')
    ws.cell(row=summary_row + 2, column=1, value=f'Range: {parsed[0][0]}\u2013{parsed[-1][0]}' if parsed else 'No invoices')
    ws.cell(row=summary_row + 3, column=1,
            value=f'Gaps found: {gaps_found}' if gaps_found else 'No gaps \u2014 fully sequential').font = (
        Font(name='Arial', bold=True, color='C00000') if gaps_found else Font(name='Arial', bold=True, color='0E6B52'))
    if unparsed:
        ws.cell(row=summary_row + 4, column=1,
                value=f'{len(unparsed)} invoice number(s) could not be parsed for sequence and need manual check: '
                      + ', '.join(str(u) for u in unparsed[:10])).font = WARN_FONT

    return wb


# ---------------- File 3: Reconciliation (workings vs Books-recorded totals) -
def build_reconciliation_workbook(period_label, std_sales, zero_export, purch_bills, rcm, cns):
    """
    Compares the workings (line-item classification used to build the return)
    against the header-level sub_total/total Zoho Books itself returns on
    each transaction — captured directly off the API record during
    classification, transaction by transaction. This catches classification
    or extraction errors without needing a separately downloaded ledger.
    """
    wb = openpyxl.Workbook()

    # --- Sales tab ---
    ws = wb.active
    ws.title = 'Sales Reconciliation'
    rows = []
    mismatches_sales = 0
    all_sales = {**{k: (v, 'Standard 15%') for k, v in std_sales.items()},
                 **{k: (v, 'Non-GCC / 0%') for k, v in zero_export.items()}}
    for inv_no, (d, treatment) in sorted(all_sales.items()):
        workings_total = round(d['taxable'] + d['vat'], 2)
        tot = d.get('total')
        diff = round(workings_total - tot, 2) if tot else None
        flag = 'No total on record' if not tot else ('OK' if abs(diff) < 0.05 else 'DIFFERENCE')
        if flag == 'DIFFERENCE':
            mismatches_sales += 1
        rows.append([inv_no, treatment, workings_total, tot, diff, flag])
    headers = ['Invoice No', 'Treatment', 'Workings Total (incl. VAT)', 'Books Header Total', 'Difference', 'Check']
    next_row = add_table(ws, 4, headers, rows, title=f'SALES RECONCILIATION - {period_label}', title_row=2)
    apply_money_cols(ws, 5, next_row - 1, [3, 4, 5])
    _flag_differences(ws, 5, next_row - 1, 6)

    # --- Purchases tab (domestic standard-rated bills only, excl. RCM) ---
    ws = wb.create_sheet('Purchase Reconciliation')
    rows = []
    mismatches_purch = 0
    for bill_no, d in sorted(purch_bills.items()):
        workings_total = round(d['taxable'] + d['vat'], 2)
        tot = d.get('total')
        diff = round(workings_total - tot, 2) if tot else None
        flag = 'No total on record' if not tot else ('OK' if abs(diff) < 0.05 else 'DIFFERENCE')
        if flag == 'DIFFERENCE':
            mismatches_purch += 1
        rows.append([bill_no, d['vendor'], workings_total, tot, diff, flag])
    headers = ['Bill No', 'Vendor', 'Workings Total (incl. VAT)', 'Books Header Total', 'Difference', 'Check']
    next_row = add_table(ws, 4, headers, rows, title=f'PURCHASE RECONCILIATION (domestic, excl. RCM) - {period_label}', title_row=2)
    apply_money_cols(ws, 5, next_row - 1, [3, 4, 5])
    _flag_differences(ws, 5, next_row - 1, 6)

    # --- RCM tab (value only, no VAT on the bill itself) ---
    ws = wb.create_sheet('RCM Reconciliation')
    rows = []
    mismatches_rcm = 0
    for bill_no, d in sorted(rcm.items()):
        workings_value = round(d['taxable'], 2)
        tot = d.get('total')
        diff = round(workings_value - tot, 2) if tot else None
        flag = 'No total on record' if not tot else ('OK' if abs(diff) < 0.05 else 'DIFFERENCE')
        if flag == 'DIFFERENCE':
            mismatches_rcm += 1
        rows.append([bill_no, d['vendor'], workings_value, tot, diff, flag])
    headers = ['Bill No', 'Vendor', 'Workings Value', 'Books Header Total', 'Difference', 'Check']
    next_row = add_table(ws, 4, headers, rows, title=f'RCM (CYBER KNIGHT) RECONCILIATION - {period_label}', title_row=2)
    apply_money_cols(ws, 5, next_row - 1, [3, 4, 5])
    _flag_differences(ws, 5, next_row - 1, 6)

    # --- Summary tab ---
    ws = wb.create_sheet('Summary', 0)
    ws['B2'] = f'RECONCILIATION SUMMARY - {period_label}'
    ws['B2'].font = TITLE_FONT
    ws['B3'] = ('Workings totals (from the classification used to build the return) vs the sub_total/total '
                'Zoho Books itself returned on each transaction, transaction by transaction. "No total on '
                'record" means that field was missing or zero on the API response for that transaction.')
    ws['B3'].font = NOTE_FONT
    total_checked = len(all_sales) + len(purch_bills) + len(rcm)
    total_mismatches = mismatches_sales + mismatches_purch + mismatches_rcm
    summary_rows = [
        ['Sales invoices checked', len(all_sales), f'{mismatches_sales} difference(s)' if mismatches_sales else 'All match'],
        ['Domestic purchase bills checked', len(purch_bills), f'{mismatches_purch} difference(s)' if mismatches_purch else 'All match'],
        ['RCM (Cyber Knight) bills checked', len(rcm), f'{mismatches_rcm} difference(s)' if mismatches_rcm else 'All match'],
        ['Total transactions checked', total_checked, f'{total_mismatches} difference(s) total' if total_mismatches else 'All match'],
    ]
    headers = ['Check', 'Count', 'Result']
    add_table(ws, 6, headers, summary_rows)
    result_font = Font(name='Arial', bold=True, color='C00000' if total_mismatches else '0E6B52')
    ws.cell(row=6 + len(summary_rows) + 2, column=2,
            value='DIFFERENCES FOUND — see the detail tabs' if total_mismatches
            else 'No differences — workings tie out exactly to Books-recorded totals').font = result_font

    return wb, total_mismatches


def _flag_differences(ws, first_row, last_row, check_col):
    red = Font(name='Arial', bold=True, color='C00000')
    for r in range(first_row, last_row + 1):
        if ws.cell(row=r, column=check_col).value == 'DIFFERENCE':
            for c in range(1, check_col + 1):
                ws.cell(row=r, column=c).font = red


# ---------------- File 7: Journal Entry (post-filing, values only) -----------
def build_journal_entry_workbook(period_label, source_filename, net_vat_due, sales_vat, purch_vat_total, rcm_vat):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Journal Entry - Post Filing'
    ws['B2'] = f'VAT SETTLEMENT JOURNAL ENTRY - {period_label} (complete after filing confirmation)'
    ws['B2'].font = TITLE_FONT
    ws['B3'] = ('Human review required: confirm the actually FILED figures below (they may differ from this '
                'computation if ZATCA amendments were made) before posting.')
    ws['B3'].font = WARN_FONT
    ws['B4'] = f'Source: values pulled from {source_filename}, sheet "VAT Return", at time of generation.'
    ws['B4'].font = NOTE_FONT

    output_total = round(sales_vat + rcm_vat, 2)
    input_total = round(purch_vat_total, 2)  # purch_vat_total already includes RCM input side

    rows = [
        ['Output VAT Payable (control a/c)', output_total, None, 'Clears output VAT recognised during the period, incl. RCM self-charge'],
        ['Input VAT Recoverable (control a/c)', None, input_total, 'Clears input VAT recognised during the period, incl. RCM'],
        ['VAT Payable to ZATCA', None, round(net_vat_due, 2), 'Net amount due per filed return \u2014 CONFIRM against actual filing before posting'],
    ]
    headers = ['Account', 'Debit', 'Credit', 'Notes']
    next_row = add_table(ws, 6, headers, rows)
    apply_money_cols(ws, 7, next_row - 1, [2, 3])
    total_row = next_row + 1
    write_total_row(ws, total_row, 1, 'Total', [2, 3], 7, next_row - 1)
    return wb


# ---------------- Orchestration ----------------------------------------------
def recalc(path):
    result = subprocess.run(['python3', RECALC_SCRIPT, path, '90'], capture_output=True, text=True)
    return result.stdout


def run(output_root, year, month, org_id=None):
    month_name = datetime(int(year), int(month), 1).strftime('%B')
    period_label = f"{month_name} {year}"
    folder_name = f"VAT - {month_name}'{str(year)[2:]}"
    out_dir = os.path.join(output_root, folder_name)
    os.makedirs(out_dir, exist_ok=True)

    invoice_records = zoho_books_api.get_invoices_for_period(year, month)
    cn_records = zoho_books_api.get_creditnotes_for_period(year, month)
    bill_records = zoho_books_api.get_bills_for_period(year, month)
    expense_records = zoho_books_api.get_expenses_for_period(year, month)

    std_sales, zero_export = extract_sales(invoice_records)
    cns = extract_credit_notes(cn_records)
    purch_bills, rcm = extract_bills(bill_records)
    purch_exp = extract_expenses(expense_records)

    files = {}

    # 1. Main return
    wb, final_row = build_return_workbook(period_label, std_sales, zero_export, cns, purch_bills, rcm, purch_exp)
    return_filename = f"Knights_Telecom_VAT_Return_{month_name}_{year}.xlsx"
    return_path = os.path.join(out_dir, return_filename)
    wb.save(return_path)
    recalc(return_path)
    wb2 = openpyxl.load_workbook(return_path, data_only=True)
    net_vat_due = wb2['VAT Return'].cell(row=final_row, column=9).value
    files['vat_return'] = return_filename

    # 2. Invoice series check
    wb = build_invoice_series_workbook(period_label, std_sales, zero_export)
    fn = f"Invoice_Series_Check_{month_name}_{year}.xlsx"
    wb.save(os.path.join(out_dir, fn)); recalc(os.path.join(out_dir, fn))
    files['invoice_series_check'] = fn

    # 3. Reconciliation: workings vs Books-recorded header totals
    wb, total_mismatches = build_reconciliation_workbook(
        period_label, std_sales, zero_export, purch_bills, rcm, cns
    )
    fn = f"Reconciliation_{month_name}_{year}.xlsx"
    wb.save(os.path.join(out_dir, fn)); recalc(os.path.join(out_dir, fn))
    files['reconciliation'] = fn

    # 4. Journal entry (values only, sourced from the recalculated return)
    sales_vat_total = wb2['VAT Return']['I19'].value or 0
    purch_vat_incl_rcm = wb2['VAT Return']['I30'].value or 0
    rcm_vat_total = sum(d['taxable'] for d in rcm.values()) * 0.15
    wb = build_journal_entry_workbook(period_label, return_filename, net_vat_due, sales_vat_total, purch_vat_incl_rcm, rcm_vat_total)
    fn = f"Journal_Entry_{month_name}_{year}.xlsx"
    wb.save(os.path.join(out_dir, fn)); recalc(os.path.join(out_dir, fn))
    files['journal_entry'] = fn

    return {
        'folder_path': out_dir,
        'files': files,
        'net_vat_due': net_vat_due,
        'reconciliation_mismatches': total_mismatches,
    }


if __name__ == '__main__':
    import sys
    out_root = sys.argv[1] if len(sys.argv) > 1 else '/mnt/user-data/outputs'
    yr = sys.argv[2] if len(sys.argv) > 2 else '2026'
    mo = sys.argv[3] if len(sys.argv) > 3 else '6'
    result = run(out_root, yr, mo)
    print(result['folder_path'])
    for k, v in result['files'].items():
        print(' -', v)
    print('Net VAT due:', result['net_vat_due'])
